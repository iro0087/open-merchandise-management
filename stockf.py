#!usr/bin python3

#this program is intended to run with frenh date format only (day/month/year)

#feel free to clean it, iknow some list are useless at this point...

from openpyxl import *

import random, openpyxl

from tkinter import *

import sys

from datetime import datetime

def quit_(envent2=None):

    sys.exit()

def stock(event3=None):

    now = datetime.now()

    day = now.strftime("%d")

    month = now.strftime("%m")

    year = now.strftime("%y")

    date = day + "/" + month + "/" + year

    s = load_workbook("stock.xlsx")

    s2 = s["Stock CDS"]  #can be modified

    s3 = s["Sorties"]

    s4 = s["Entrées"]

    sb = load_workbook("e.xlsx")

    sb2 = sb.active

    t = 2

    stockl = []

    sortiel = []

    entreel = []

    t2 = 3

    while s2.cell(row=t2, column=2).value != None:

        t2 += 1

    t2 = t2 - 1

    while s2.cell(row=t2, column=t).value != None:

        stockl.append(float(s2.cell(row=t2, column=t).value))

        sortiel.append(0)

        entreel.append(0)

        t += 1

    t_element = t - 1

    t = 1

    while s3.cell(row=t, column=1).value != None:

        t += 1

    t_sc = t #can be modified

    t = 1

    while s4.cell(row=t, column=1).value != None:

        t += 1

    t_ec = t #can be modified

    if t_ec != sb2.cell(row=1, column=2).value:
        
        t = sb2.cell(row=1, column=2).value

        t2 = 3

        t_adjust = t2 #can be modified

        while t < t_ec:

            while t2 < t_element:

                if s4.cell(row=t, column=t2).value != None:

                    entreel[t2 - t_adjust] = entreel[t2 - t_adjust] + float(s4.cell(row=t, column=t2).value)

                t2 += 1

            t2 = 3

            t += 1

        t = 0

        while t < len(stockl):

            stockl[t] = stockl[t] + entreel[t]

            t += 1

    if t_sc != sb2.cell(row=2, column=2).value:
        
        t = sb2.cell(row=2, column=2).value

        t2 = 3

        t_adjust = t2 #can be modified

        while t < t_sc:

            while t2 < t_element:

                if s3.cell(row=t, column=t2).value != None:

                    sortiel[t2 - t_adjust] = sortiel[t2 - t_adjust] + float(s3.cell(row=t, column=t2).value)

                t2 += 1

            t2 = 3

            t += 1

        t = 0

        while t < len(stockl):

            stockl[t] = stockl[t] - sortiel[t]

            t += 1

    sb2.cell(row=1, column=2).value = t_ec 

    sb2.cell(row=2, column=2).value = t_sc

    t_actl = 1

    while s2.cell(row=t_actl, column=1).value != None:

        t_actl += 1

    s2.cell(row=t_actl, column=1).value = str(date)

    t = 2

    while t - 2 < len(stockl):

        s2.cell(row=t_actl, column=t).value = stockl[t - 2]

        t += 1

    sb.save("e.xlsx")

    s.save("stock.xlsx")

def searchw(event=None):

    def search(event7=None):

        var = varf.get()

        vard_ = str(vard.get())

        vard2_ = str(vard2.get())

        t = 0

        ld = []

        vl = []

        while t < len(vard_):

            if vard_[t] == "/" or t + 1 == len(vard_):

                if t + 1 == len(vard_):

                    ld.append(vard_[t])

                j = int("".join(ld))

                vl.append(j)

                ld = []

            else:

                ld.append(vard_[t])

            t += 1

        t = 0

        while t < len(vard2_):

            if vard2_[t] == "/" or t + 1 == len(vard2_):

                if t + 1 == len(vard2_):

                    ld.append(vard2_[t])

                j = int("".join(ld))

                vl.append(j)

                ld = []

            else:

                ld.append(vard2_[t])

            t += 1

        s = load_workbook("stock.xlsx")

        if var == "s":

            s2 = s["Stock CDS"]

        if var == "r":

            s2 = s["Entrées"]

        if var == "d" or var == "o":

            s2 = s["Sorties"]

        t = 4

        ld = []

        ld2 = []

        ldf1 = []

        ldf2 = []

        ldf3 = []

        t_count = 0

        while s2.cell(row=t, column=2).value != None:

            ld.append(s2.cell(row=t, column=2).value)

            t += 1

        t = 0

        t2 = 0

        t_adjust = 4

        while t < len(ld):

            while t2 < len(ld[t]):

                if ld[t][t2] == "/" or t2 + 1 == len(ld[t]):

                    if t2 + 1 == len(ld[t]):

                        ld2.append(ld[t][-1])

                    j = int("".join(ld2))

                    ld2 = []

                    if t_count == 0:

                        ldf1.append(j)

                    if t_count == 1:

                        ldf2.append(j)

                    if t_count == 2:

                        ldf3.append(j)

                    t_count += 1

                else:

                    ld2.append(ld[t][t2])

                t2 += 1

            t2 = 0

            t_count = 0

            ld2 = []

            t += 1

        ldf4 = [ldf3, ldf2, ldf1]

        t = 0

        lix1 = []

        lix2 = []

        lix3 = []

        lixf = [lix3, lix2, lix1]

        while t < len(ldf4):

            while t2 < len(ldf4[t]):

                lixf[t].append(abs(vl[2 - t] - ldf4[t][t2]))

                t2 += 1

            t2 = 0

            t += 1

        t = 0

        somma_l = []

        somma_l2 = []

        somma_l3 = []

        somma_lf = [somma_l3, somma_l2, somma_l]

        while t < len(lixf[0]):

            somma_l.append(lix1[t]) #day diff

            somma_l2.append(lix2[t]) #month diff

            somma_l3.append(lix3[t]) #year diff

            t += 1

        t = 0

        t2 = 0

        maxo = max(somma_l2)

        while t < len(somma_lf) - 1:

            t2 = somma_lf[t].index(min(somma_lf[t]))

            tmin = t2 + 1

            t3 = 0

            t_adjust = 0

            if t == 0:

                while t3 < t2:

                    somma_l2[t3] = maxo

                    t3 += 1

            while min(somma_lf[t]) == somma_lf[t][t2] and t2 + 1 < len(somma_lf[t]):

                t2 += 1

            tmax = t2

            t3 = tmax

            if t == 0:

                while t3 < len(somma_l2) - 1:

                    somma_l2[t3] = maxo

                    t3 += 1

            t += 1
        
        t = tmin - 1

        day = []

        if len(day) == 0:

            day.append(tmin)

        else:

            while t < tmax: 

                day.append(somma_l[t])

                t += 1

        index1 = day.index(min(day)) + 3 + tmin

        if vl[2] < ldf3[index1 - 3 - tmin]:

            index1 = 4

        #condition à develpper concernant les mois

        t = 0

        ldf4 = [ldf3, ldf2, ldf1]

        t = 0

        t2 = 0

        lix1 = []

        lix2 = []

        lix3 = []

        lixf = [lix3, lix2, lix1]

        while t < len(ldf4):

            while t2 < len(ldf4[t]):

                lixf[t].append(abs(vl[len(vl) - 1 - t] - ldf4[t][t2]))

                t2 += 1

            t2 = 0

            t += 1

        t = 0

        somma_l = []

        somma_l2 = []

        somma_l3 = []

        somma_lf = [somma_l3, somma_l2, somma_l]

        while t < len(lixf[0]):

            somma_l.append(lix1[t]) #day diff

            somma_l2.append(lix2[t]) #month diff

            somma_l3.append(lix3[t]) #year diff

            t += 1

        t = 0

        t2 = 0

        maxo = max(somma_l2)

        while t < len(somma_lf) - 1:

            t2 = somma_lf[t].index(min(somma_lf[t]))

            tmin = t2 + 1

            t3 = 0

            t_adjust = 0

            if t == 0:

                while t3 < t2:

                    somma_l2[t3] = maxo

                    t3 += 1

            while min(somma_lf[t]) == somma_lf[t][t2] and t2 + 1 < len(somma_lf[t]):

                t2 += 1

            tmax = t2

            t3 = tmax

            if t == 0:

                while t3 < len(somma_l2) - 1:

                    somma_l2[t3] = maxo

                    t3 += 1

            t += 1
        
        t = tmin - 1

        day = []

        if len(day) == 0:

            day.append(tmin)

        else:

            while t < tmax: 

                day.append(somma_l[t])

                t += 1

        index2 = day.index(min(day)) + 3 + tmin

        if vl[-1] < ldf3[index2 - 3 - tmin]:

            index2 = 4

        t = 3 #could be changed

        t2 = 0

        ele_l = []

        ded = load_workbook("dedicated.xlsx")

        ded2 = ded.active

        ded2.cell(row=1, column=1).value += 1

        name = "results" + str(ded2.cell(row=1, column=1).value) + ".xlsx"

        ded.save("dedicated.xlsx")

        sf = openpyxl.Workbook(name)

        sf.save(name)

        sf = load_workbook(name)

        sf2 = sf.active

        while s2.cell(row=2, column=t).value != None:

            sf2.cell(row=1, column=t-2).value = str(s2.cell(row=2, column=t).value)

            ele_l.append(0)

            t += 1

        t = 3

        while index1 <= index2:

            while t < len(ele_l):

                if s2.cell(row=index1, column=t).value != None:

                    ele_l[t - 3] += int(s2.cell(row=index1, column=t).value)

                t += 1

            t = 3

            index1 += 1

        t = 1

        while t < len(ele_l):

            sf2.cell(row=2, column=t).value = ele_l[t - 1]

            t += 1

        sf.save(name)

        Label(fen2, text="").pack()

        Label(fen2, text="done").pack()

    fen2 = Tk()

    fen2.geometry("400x350")

    varf = Entry(fen2, width=30)

    varf.focus_set()

    def varf_(event6=None):

        varf.focus_set()

    vard = Entry(fen2, width=30)

    def vard_(event4=None):

        vard.focus_set()

    vard2 = Entry(fen2, width=30)

    def vard2_(event5=None):

        vard2.focus_set()
    
    Label(fen2, text="").pack()
    
    Label(fen2, text="Search for stock, received or delivered/orders (s, r, d/o)").pack()

    Label(fen2, text="").pack()

    varf.pack()

    Label(fen2, text="").pack()

    Label(fen2, text="Begining").pack()

    Label(fen2, text="").pack()

    vard.pack()

    Label(fen2, text="").pack()

    Label(fen2, text="End").pack()

    Label(fen2, text="").pack()

    vard2.pack()

    Label(fen2, text="").pack()

    Button(fen2, text="PROCEED", command=search, bg="yellow").pack()

    fen2.bind("w", searchw)

    fen2.bind("q", quit_)

    fen2.bind("a", stock)

    #fen2.bind("<Left>", vard_)

    #fen2.bind("<Right>", vard2_)

    fen2.bind("<Down>", varf_)

    fen2.mainloop()

fen = Tk()

fen.geometry("240x190")

Label(fen, text="").pack()

Button(fen, text="REFRESH", command=stock, bg="yellow").pack()

Label(fen, text="").pack()

Button(fen, text="Search between date", command=searchw, bg="yellow").pack()

fen.bind("w", searchw)

fen.bind("q", quit_)

fen.bind("a", stock)

Label(fen, text="").pack()

Button(fen, text="TERMINATE", command=quit_, bg="red").pack()

fen.mainloop()




