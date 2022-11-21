#!usr/bin python3

#this program is intended to run with frenh date format only (day/month/year)

from openpyxl import load_workbook

import random

from tkinter import *

import sys

import time

def quit_(envent2=None):

    sys.exit()

def stock(event3=None):

    s = load_workbook("stock.xlsx")

    s2 = s["Stock CDS"]  #can be modified

    s3 = s["Sorties"]

    s4 = s["Entrées"]

    sb = load_workbook("e.xlsx")

    sb2 = sb.active

    t = 2

    stockl = []

    sortiel = []

    entreel = []

    while s2.cell(row=3, column=t).value != None:

        stockl.append(float(s2.cell(row=3, column=t).value))

        sortiel.append(0)

        entreel.append(0)

        t += 1

    t_element = t - 1

    t = 1

    while s3.cell(row=t, column=1).value != None:

        t += 1

    t_sc = t - 4 #can be modified

    t = 1

    while s4.cell(row=t, column=1).value != None:

        t += 1

    t_ec = t - 4 #can be modified

    if t_ec != sb2.cell(row=1, column=2).value:
        
        t = t_ec 

        t2 = 3

        t_adjust = t2 #can be modified

        while t < t_sc:

            while t2 < t_element:

                if s4.cell(row=t, column=t2).value != None:

                    s4.cell(row=t, column=t2).value = float(s4.cell(row=t, column=t2).value)

                    entreel[t2 - t_adjust] = entreel[t2 - t_adjust] + float(s4.cell(row=t, column=t2).value)

                t2 += 1

            t2 = 3

            t += 1

        t = 0

        while t < len(stockl):

            stockl[t] = stockl[t] + entreel[t]

            t += 1

    if t_sc != sb2.cell(row=2, column=2).value:
        
        t = t_sc 

        t2 = 3

        t_adjust = t2 #can be modified

        while t < t_sc:

            while t2 < t_element:

                if s3.cell(row=t, column=t2).value != None:

                    s3.cell(row=t, column=t2).value = float(s3.cell(row=t, column=t2).value)

                    sortiel[t2 - t_adjust] = sortiel[t2 - t_adjust] + float(s3.cell(row=t, column=t2).value)

                t2 += 1

            t2 = 3

            t += 1

        t = 0

        while t < len(stockl):

            stockl[t] = stockl[t] - sortiel[t]

            t += 1

    sb2.cell(row=1, column=2).value = t_ec

    sb2.cell(row=2, column=2).value = t_sc

    sb.save("e.xlsx")

    s.save("stock.xlsx")

def searchw(event=None):

    def search(event7=None):

        var = varf.get()

        vard_ = str(vard.get())

        vard2_ = str(vard2.get())

        t = 0

        ld = []

        vl = []

        while t < len(vard_):

            if vard_[t] == "/" or t + 1 == len(vard_):

                if t + 1 == len(vard_):

                    ld.append(vard_[t])

                j = int("".join(ld))

                vl.append(j)

                ld = []

            else:

                ld.append(vard_[t])

            t += 1

        t = 0

        while t < len(vard2_):

            if vard2_[t] == "/" or t + 1 == len(vard2_):

                if t + 1 == len(vard2_):

                    ld.append(vard2_[t])

                j = int("".join(ld))

                vl.append(j)

                ld = []

            else:

                ld.append(vard2_[t])

            t += 1

        s = load_workbook("stock.xlsx")

        sbb = load_workbook("output.xlsx")

        sbb2 = sbb.active

        if var == "s":

            s2 = s["Stock CDS"]

        if var == "r":

            s2 = s["Entrées"]

        if var == "d" or var == "o":

            s2 = s["Sorties"]

        t = 4

        ld = []

        ld2 = []

        ldf1 = []

        ldf2 = []

        ldf3 = []

        t_count = 0

        while s2.cell(row=t, column=2).value != None:

            ld.append(s2.cell(row=t, column=2).value)

            t += 1

        t = 0

        t2 = 0

        t_adjust = 4

        while t < len(ld):

            while t2 < len(ld[t]):

                if ld[t][t2] == "/" or t2 + 1 == len(ld[t]):

                    if t2 + 1 == len(ld[t]):

                        ld2.append(ld[t][-1])

                    j = int("".join(ld2))

                    ld2 = []

                    if t_count == 0:

                        ldf1.append(j)

                    if t_count == 1:

                        ldf2.append(j)

                    if t_count == 2:

                        ldf3.append(j)

                    t_count += 1

                else:

                    ld2.append(ld[t][t2])

                t2 += 1

            t2 = 0

            t_count = 0

            ld2 = []

            t += 1

        ldf4 = [ldf3, ldf2, ldf1]

        t = 0

        index1 = 0

        index2 = 0

        lix1 = []

        lix2 = []

        lix3 = []

        lixf = [lix3, lix2, lix1]

        print(ldf4)

        print("vl", vl)

        while t < len(ldf4):

            while t2 < len(ldf4[t]):

                lixf[t].append(abs(vl[2 - t] - ldf4[t][t2]))

                t2 += 1

            t2 = 0

            t += 1

        t = 0

        somma_l = []

        somma_l2 = []

        somma_l3 = []

        somma_lf = [somma_l3, somma_l2, somma_l]

        while t < len(lixf[0]):

            somma_l.append(lix1[t]) #day diff

            somma_l2.append(lix2[t]) #month diff

            somma_l3.append(lix3[t]) #year diff

            t += 1

        t = 0

        t2 = 0

        tmaxl = []

        tminl = []

        maxo = max(somma_l2)

        while t < len(somma_lf) - 1:

            tminl.append(somma_lf[t].index(min(somma_lf[t])) + 1)

            t2 = tminl[-1]

            print(t2)

            while somma_lf[t][t2] == min(somma_lf[t]) and t2 + 1 < len(somma_lf[t]):

                t2 += 1

            tmaxl.append(t2)

            t2 = 0

            if t == 0:

                while t2 < tminl[-1]:

                    somma_l2[t2] = maxo

                    t2 += 1

                t2 = tmaxl[-1]

                while t2 < len(somma_l2):

                    somma_l2[t2] = maxo

                    t2 += 1

                t2 = 0

            t += 1

        t = 0

        maxo = max(somma_l)

        while t < tminl[-1]:

            somma_l[t] = maxo

            t += 1

        t = tmaxl[-1]

        while t < len(somma_l):

            somma_l[t] = maxo

            t += 1

        t = tminl[-1]

        tminl = [-1]

        jour = []

        while t < tmaxl[-1]:

            jour.append(somma_l[t])

            t += 1

        u = jour.index(min(jour)) + tminl[-1]

        print(u)

        print(tminl, tmaxl)

    fen2 = Tk()

    fen2.bind("<F2>", search)

    fen2.geometry("400x320")

    varf = Entry(fen2, width=30)

    varf.focus_set()

    def varf_(event6=None):

        varf.focus_set()

    vard = Entry(fen2, width=30)

    def vard_(event4=None):

        vard.focus_set()

    vard2 = Entry(fen2, width=30)

    def vard2_(event5=None):

        vard2.focus_set()
    
    Label(fen2, text="").pack()
    
    Label(fen2, text="Search for stock, received or delivered/orders (s, r, d/o)").pack()

    Label(fen2, text="").pack()

    varf.pack()

    Label(fen2, text="").pack()

    Label(fen2, text="Begining").pack()

    Label(fen2, text="").pack()

    vard.pack()

    Label(fen2, text="").pack()

    Label(fen2, text="End").pack()

    Label(fen2, text="").pack()

    vard2.pack()

    Label(fen2, text="").pack()

    Button(fen2, text="PROCEED", command=search, bg="yellow").pack()

    fen2.bind("w", searchw)

    fen2.bind("q", quit_)

    fen2.bind("a", stock)

    fen2.bind("<Left>", vard_)

    fen2.bind("<Right>", vard2_)

    fen2.bind("<Down>", varf_)

    fen2.mainloop()

fen = Tk()

fen.geometry("240x190")

Label(fen, text="").pack()

Button(fen, text="REFRESH", command=stock, bg="yellow").pack()

Label(fen, text="").pack()

Button(fen, text="Search between date", command=searchw, bg="yellow").pack()

fen.bind("w", searchw)

fen.bind("q", quit_)

fen.bind("a", stock)

Label(fen, text="").pack()

Button(fen, text="TERMINATE", command=quit_, bg="red").pack()

fen.mainloop()




