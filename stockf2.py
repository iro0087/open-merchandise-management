#!usr/bin python3

#this program is intended to run with frenh date format only (day/month/year)

#feel free to clean it, iknow some list are useless at this point...

from openpyxl import *

import random, openpyxl

from tkinter import *

import sys

from datetime import datetime

def quit_(envent2=None):

    sys.exit()

def stock(event3=None):

    now = datetime.now()

    day = now.strftime("%d")

    month = now.strftime("%m")

    year = now.strftime("%y")

    date = day + "/" + month + "/" + year

    s = load_workbook("stock.xlsx")

    s2 = s["Stock"]  #can be modified

    s3 = s["Out"]

    s4 = s["In"]

    sb = load_workbook("e.xlsx")

    sb2 = sb.active

    t = 2

    stockl = []

    sortiel = []

    entreel = []

    t2 = 3

    while s2.cell(row=t2, column=2).value != None:

        t2 += 1

    t2 = t2 - 1

    while s2.cell(row=t2, column=t).value != None:

        stockl.append(float(s2.cell(row=t2, column=t).value))

        sortiel.append(0)

        entreel.append(0)

        t += 1

    t_element = t - 1

    t = 1

    while s3.cell(row=t, column=1).value != None:

        t += 1

    t_sc = t #can be modified

    t = 1

    while s4.cell(row=t, column=1).value != None:

        t += 1

    t_ec = t #can be modified

    if t_ec != sb2.cell(row=1, column=2).value:
        
        t = sb2.cell(row=1, column=2).value

        t2 = 3

        t_adjust = t2 #can be modified

        while t < t_ec:

            while t2 < t_element:

                if s4.cell(row=t, column=t2).value != None:

                    entreel[t2 - t_adjust] = entreel[t2 - t_adjust] + float(s4.cell(row=t, column=t2).value)

                t2 += 1

            t2 = 3

            t += 1

        t = 0

        while t < len(stockl):

            stockl[t] = stockl[t] + entreel[t]

            t += 1

    if t_sc != sb2.cell(row=2, column=2).value:
        
        t = sb2.cell(row=2, column=2).value

        t2 = 3

        t_adjust = t2 #can be modified

        while t < t_sc:

            while t2 < t_element:

                if s3.cell(row=t, column=t2).value != None:

                    sortiel[t2 - t_adjust] = sortiel[t2 - t_adjust] + float(s3.cell(row=t, column=t2).value)

                t2 += 1

            t2 = 3

            t += 1

        t = 0

        while t < len(stockl):

            stockl[t] = stockl[t] - sortiel[t]

            t += 1

    sb2.cell(row=1, column=2).value = t_ec 

    sb2.cell(row=2, column=2).value = t_sc

    t_actl = 1

    while s2.cell(row=t_actl, column=1).value != None:

        t_actl += 1

    s2.cell(row=t_actl, column=1).value = str(date)

    t = 2

    while t - 2 < len(stockl):

        s2.cell(row=t_actl, column=t).value = stockl[t - 2]

        t += 1

    sb.save("e.xlsx")

    s.save("stock.xlsx")

    Label(fen, text="").pack()

    Label(fen, text="done").pack()

def searchw(event=None):

    def search(event7=None):

        ded = load_workbook("dedicated.xlsx")

        ded2 = ded.active

        ded2.cell(row=1, column=1).value += 1

        name = "results" + str(ded2.cell(row=1, column=1).value) + ".xlsx"

        ded.save("dedicated.xlsx")

        sf = openpyxl.Workbook(name)

        sf.save(name)

        sf = load_workbook(name)

        sf2 = sf.active

        s = load_workbook("stock.xlsx")

        var = varf.get()

        vard_ = str(vard.get())

        vard2_ = str(vard2.get())

        t = 0

        ld = []

        vl = []

        while t < len(vard_):

            if vard_[t] == "/" or t + 1 == len(vard_):

                if t + 1 == len(vard_):

                    ld.append(vard_[t])

                j = int("".join(ld))

                vl.append(j)

                ld = []

            else:

                ld.append(vard_[t])

            t += 1

        t = 0

        while t < len(vard2_):

            if vard2_[t] == "/" or t + 1 == len(vard2_):

                if t + 1 == len(vard2_):

                    ld.append(vard2_[t])

                j = int("".join(ld))

                vl.append(j)

                ld = []

            else:

                ld.append(vard2_[t])

            t += 1

        if var == "s":

            s2 = s["Stock"]

        if var == "r":

            s2 = s["In"]

        if var == "d" or var == "o":

            s2 = s["Out"]

        t = 4

        t2 = 0

        al = []

        dl = []

        rl = []

        ml = []

        yl = []

        while s2.cell(row=t, column=2).value != None:

            al.append(s2.cell(row=t, column=2).value)

            t2 = 0

            while t2 < 2:

                rl.append(al[-1][t2])

                t2 += 1

            t2 = 3

            rj = "".join(rl)

            dl.append(int(rj))

            rl = []

            while t2 < 5:

                rl.append(al[-1][t2])

                t2 += 1

            t2 = 6

            rj = "".join(rl)

            ml.append(int(rj))

            rl = []

            while t2 < len(s2.cell(row=t, column=2).value):

                rl.append(al[-1][t2])

                t2 += 1

            rj = "".join(rl)

            yl.append(int(rj))

            rl = []

            t += 1

        fl = [yl, ml, dl]

        tl = []

        T = 0

        T2 = 0

        p = 2

        t = 0

        t2 = 0

        i1 = []

        i2 = []

        dd = []

        dm = []

        dy = []

        dd2 = []

        dm2 = []

        dy2 = []

        tl = [dy, dm, dd, dy2, dm2, dd2]

        il = []

        while T < 2:

            while t < 3:

                while t2 < len(fl[t]):

                    tl[T2].append(abs(fl[t][t2] - vl[p - t]))

                    t2 += 1

                t2 = 0

                T2 += 1

                t += 1

            p = 5

            t = 0

            T += 1

        t = 0

        t2 = 0

        trn = 0

        while t < len(tl): #this algorythm finds the closest date to the input

            mn = min(tl[t])

            il.append(tl[t].index(mn))

            while tl[t].count(mn) > 0:

                tl[t].remove(min(tl[t]))

                t2 += 1

            il.append(il[-1] + t2 - 1)

            t2 = 0

            if t + 1 < len(tl) and t != 2:

                mx = max(tl[t + 1])

                while t2 < il[-2]:

                    tl[t + 1][t2] = mx + 1

                    t2 += 1

                t2 = il[-1] + 1

                while t2 < len(tl[t + 1]):

                    tl[t + 1][t2] = mx + 1 

                    t2 += 1

                t2 = 0

            trn = 0

            t += 1

        t = 3  #end of the algorythm

        lt = []

        while s2.cell(row=2, column=t).value != None:

            lt.append(0)

            t += 1

        t = il[4] + 4  #takes the largest interval

        t2 = 3

        while t < il[-1] + 5:

            while t2 < len(lt) + 3:

                if s2.cell(row=t, column=t2).value != None:

                    lt[t2 - 3] = lt[t2 - 3] + s2.cell(row=t, column=t2).value

                t2 += 1

            t2 = 3

            t += 1

        t = 3

        while s2.cell(row=2, column=t).value != None:

            sf2.cell(row=1, column=t-2).value = s2.cell(row=2, column=t).value

            sf2.cell(row=2, column=t-2).value = lt[t - 3]

            t += 1

        sf.save(name)

        sen1 = "The first date is between" + str(s2.cell(row=il[4] + 3, column=2).value) \
                + " and " + str(s2.cell(row=il[4] + 5, column=2).value)

        sen2 = "The second date is between" + str(s2.cell(row=il[-1] + 3, column=2).value) \
                + " and " + str(s2.cell(row=il[-1] + 5, column=2).value)

        Label(fen2, text="").pack()

        Label(fen2, text="done").pack()

        Label(fen2, text="").pack()

        Label(fen2, text=sen1).pack()

        Label(fen2, text="").pack()

        Label(fen2, text=sen2).pack()

    fen2 = Tk()

    varf = Entry(fen2, width=30)

    varf.focus_set()

    def varf_(event6=None):

        varf.focus_set()

    vard = Entry(fen2, width=30)

    def vard_(event4=None):

        vard.focus_set()

    vard2 = Entry(fen2, width=30)

    def vard2_(event5=None):

        vard2.focus_set()
    
    Label(fen2, text="").pack()
    
    Label(fen2, text="Search for stock, received or delivered/orders (s, r, d/o)").pack()

    Label(fen2, text="").pack()

    varf.pack()

    Label(fen2, text="").pack()

    Label(fen2, text="Begining").pack()

    Label(fen2, text="").pack()

    vard.pack()

    Label(fen2, text="").pack()

    Label(fen2, text="End").pack()

    Label(fen2, text="").pack()

    vard2.pack()

    Label(fen2, text="").pack()

    Button(fen2, text="PROCEED", command=search, bg="yellow").pack()

    fen2.bind("w", searchw)

    fen2.bind("q", quit_)

    fen2.bind("a", stock)

    #fen2.bind("<Left>", vard_)

    #fen2.bind("<Right>", vard2_)

    fen2.bind("<Down>", varf_)

    fen2.mainloop()

fen = Tk()

fen.geometry("240x190")

Label(fen, text="").pack()

Button(fen, text="REFRESH", command=stock, bg="yellow").pack()

Label(fen, text="").pack()

Button(fen, text="Search between date", command=searchw, bg="yellow").pack()

fen.bind("w", searchw)

fen.bind("q", quit_)

fen.bind("a", stock)

Label(fen, text="").pack()

Button(fen, text="TERMINATE", command=quit_, bg="red").pack()

fen.mainloop()




